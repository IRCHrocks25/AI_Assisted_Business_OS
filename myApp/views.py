from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.utils import timezone
import json
import openai
import logging

logger = logging.getLogger(__name__)
import cloudinary
import cloudinary.uploader
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .models import ExtractionTemplate, Document, ExtractionRun, CRMRecord, DemoEmail, DemoEmailResult, EmailCRMRecord, DemoConversation, DemoMessage, DemoTicket, DemoLead, DemoQualification, DemoEvent, DemoWorkflowRun, QualityConversation, DemoRiskAnalysis, DemoAnalyticsSnapshot, DemoInsight, CRMContact, CRMTag, CRMTask, CRMTicket, CRMActivity, CRMAutomationRun
from .extraction_utils import extract_text_from_pdf, extract_with_llm, validate_extraction
import requests

def home(request):
    return render(request, 'index.html')

def sop_bot_demo(request):
    return render(request, 'sop_bot_demo.html')

def booking_bot_demo(request):
    return render(request, 'booking_bot_demo.html')

def voice_agent_demo(request):
    return render(request, 'voice_agent_demo.html')

def custom_ai_systems(request):
    return render(request, 'custom_ai_systems.html')

def geo_demo(request):
    return render(request, 'geo_demo.html')

def email_automation_demo(request):
    """Email automation demo page"""
    emails = DemoEmail.objects.all()
    return render(request, 'email_automation_demo.html', {'emails': emails})

@csrf_exempt
def process_email_api(request):
    """Process email with AI - returns analysis and draft reply"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        email_id = data.get('email_id')
        
        if not email_id:
            return JsonResponse({'error': 'email_id required'}, status=400)
        
        email = get_object_or_404(DemoEmail, id=email_id)
        
        # Check if result already exists
        result = None
        if hasattr(email, 'ai_result'):
            result = email.ai_result
        else:
            # Process with AI (using OpenAI)
            client = openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
            
            prompt = f"""Analyze this email and provide a structured response:

Subject: {email.subject}
From: {email.from_name} <{email.from_email}>
Body: {email.body}

Provide analysis in this exact JSON format:
{{
    "intent": "one of: Sales Inquiry, Support Request, Partnership, Complaint, General Question",
    "urgency": "one of: Low, Medium, High",
    "sentiment": "one of: Positive, Neutral, Negative",
    "suggested_owner": "Team name (e.g., Sales Team, Support Team, etc.)",
    "crm_action": "Action to take (e.g., Create Opportunity, Create Ticket, Escalate, etc.)",
    "draft_reply": "A professional, helpful draft reply (2-3 paragraphs)",
    "follow_up_hours": 48,
    "confidence_score": 85
}}"""

            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are an AI email assistant. Analyze emails and provide structured responses in JSON format only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7,
                max_tokens=800
            )
            
            content = response.choices[0].message.content.strip()
            
            # Extract JSON from response (handle markdown code blocks)
            if '```json' in content:
                content = content.split('```json')[1].split('```')[0].strip()
            elif '```' in content:
                content = content.split('```')[1].split('```')[0].strip()
            
            ai_data = json.loads(content)
            
            # Create result
            result = DemoEmailResult.objects.create(
                email=email,
                intent=ai_data.get('intent', 'General Question'),
                urgency=ai_data.get('urgency', 'Medium'),
                sentiment=ai_data.get('sentiment', 'Neutral'),
                suggested_owner=ai_data.get('suggested_owner', 'General Team'),
                crm_action=ai_data.get('crm_action', 'Create Task'),
                draft_reply=ai_data.get('draft_reply', 'Thank you for your email. We will review and respond shortly.'),
                follow_up_hours=ai_data.get('follow_up_hours', 48),
                confidence_score=ai_data.get('confidence_score', 85)
            )
            
            email.is_processed = True
            email.save()
        
        # Check if CRM record already exists for this email
        existing_crm_record = EmailCRMRecord.objects.filter(email_id=str(email.id)).first()
        
        if not existing_crm_record:
            # Generate IDs for CRM, Task, and Ticket
            import random
            crm_record_id = f"CRM-{random.randint(100000, 999999)}"
            task_id = f"TASK-{random.randint(100000, 999999)}"
            ticket_id = f"TKT-{random.randint(100000, 999999)}"
            
            # Create CRM record
            EmailCRMRecord.objects.create(
                email_id=str(email.id),
                email_subject=email.subject,
                email_from=email.from_email,
                email_from_name=email.from_name,
                intent=result.intent,
                urgency=result.urgency,
                sentiment=result.sentiment,
                assigned_team=result.suggested_owner,
                crm_action=result.crm_action,
                task_id=task_id,
                ticket_id=ticket_id,
                crm_record_id=crm_record_id,
                follow_up_hours=result.follow_up_hours
            )
            crm_record_id_used = crm_record_id
            task_id_used = task_id
            ticket_id_used = ticket_id
        else:
            crm_record_id_used = existing_crm_record.crm_record_id
            task_id_used = existing_crm_record.task_id
            ticket_id_used = existing_crm_record.ticket_id
        
        return JsonResponse({
            'intent': result.intent,
            'urgency': result.urgency,
            'sentiment': result.sentiment,
            'suggested_owner': result.suggested_owner,
            'crm_action': result.crm_action,
            'draft_reply': result.draft_reply,
            'follow_up_hours': result.follow_up_hours,
            'confidence_score': result.confidence_score,
            'crm_record_id': crm_record_id_used,
            'task_id': task_id_used,
            'ticket_id': ticket_id_used,
        })
        
        # Process with AI (using OpenAI)
        client = openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        
        prompt = f"""Analyze this email and provide a structured response:

Subject: {email.subject}
From: {email.from_name} <{email.from_email}>
Body: {email.body}

Provide analysis in this exact JSON format:
{{
    "intent": "one of: Sales Inquiry, Support Request, Partnership, Complaint, General Question",
    "urgency": "one of: Low, Medium, High",
    "sentiment": "one of: Positive, Neutral, Negative",
    "suggested_owner": "Team name (e.g., Sales Team, Support Team, etc.)",
    "crm_action": "Action to take (e.g., Create Opportunity, Create Ticket, Escalate, etc.)",
    "draft_reply": "A professional, helpful draft reply (2-3 paragraphs)",
    "follow_up_hours": 48,
    "confidence_score": 85
}}"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an AI email assistant. Analyze emails and provide structured responses in JSON format only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=800
        )
        
        content = response.choices[0].message.content.strip()
        
        # Extract JSON from response (handle markdown code blocks)
        if '```json' in content:
            content = content.split('```json')[1].split('```')[0].strip()
        elif '```' in content:
            content = content.split('```')[1].split('```')[0].strip()
        
        ai_data = json.loads(content)
        
        # Create result
        result = DemoEmailResult.objects.create(
            email=email,
            intent=ai_data.get('intent', 'General Question'),
            urgency=ai_data.get('urgency', 'Medium'),
            sentiment=ai_data.get('sentiment', 'Neutral'),
            suggested_owner=ai_data.get('suggested_owner', 'General Team'),
            crm_action=ai_data.get('crm_action', 'Create Task'),
            draft_reply=ai_data.get('draft_reply', 'Thank you for your email. We will review and respond shortly.'),
            follow_up_hours=ai_data.get('follow_up_hours', 48),
            confidence_score=ai_data.get('confidence_score', 85)
        )
        
        email.is_processed = True
        email.save()
        
        # Generate IDs for CRM, Task, and Ticket
        import random
        crm_record_id = f"CRM-{random.randint(100000, 999999)}"
        task_id = f"TASK-{random.randint(100000, 999999)}"
        ticket_id = f"TKT-{random.randint(100000, 999999)}"
        
        # Create CRM record
        EmailCRMRecord.objects.create(
            email_id=str(email.id),
            email_subject=email.subject,
            email_from=email.from_email,
            email_from_name=email.from_name,
            intent=result.intent,
            urgency=result.urgency,
            sentiment=result.sentiment,
            assigned_team=result.suggested_owner,
            crm_action=result.crm_action,
            task_id=task_id,
            ticket_id=ticket_id,
            crm_record_id=crm_record_id,
            follow_up_hours=result.follow_up_hours
        )
        
        return JsonResponse({
            'intent': result.intent,
            'urgency': result.urgency,
            'sentiment': result.sentiment,
            'suggested_owner': result.suggested_owner,
            'crm_action': result.crm_action,
            'draft_reply': result.draft_reply,
            'follow_up_hours': result.follow_up_hours,
            'confidence_score': result.confidence_score,
            'crm_record_id': crm_record_id,
            'task_id': task_id,
            'ticket_id': ticket_id,
        })
        
    except Exception as e:
        return JsonResponse({
            'error': 'Processing failed',
            'message': str(e)
        }, status=500)

def extraction_demo(request):
    """Main extraction demo page"""
    templates = ExtractionTemplate.objects.all()
    documents = Document.objects.all()
    return render(request, 'extraction_demo.html', {
        'templates': templates,
        'documents': documents
    })

@csrf_exempt
def chat_api(request):
    """Handle chatbot API requests"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        message = data.get('message', '')
        
        if not message:
            return JsonResponse({'error': 'Message is required'}, status=400)
        
        # Check if API key is configured (loaded from .env via dotenv in settings)
        api_key = settings.OPENAI_API_KEY
        if not api_key:
            logger.error('OpenAI API key not configured. Make sure OPENAI_API_KEY is set in your .env file.')
            return JsonResponse({
                'error': 'OpenAI API key not configured',
                'response': 'I apologize, but the AI service is not currently configured. Please contact support.'
            }, status=500)
        
        # Initialize OpenAI client
        client = openai.OpenAI(api_key=api_key)
        
        # System prompt for the chatbot
        system_prompt = """You are a helpful AI assistant for AI Assisted Business OS, a company that provides AI-powered business solutions including chatbots, automation, customer support, sales qualification, and various AI integrations.

Your role is to:
- Answer questions about AI Assisted Business OS services and capabilities
- Help visitors understand how AI can transform their business operations
- Provide information about the catalogue of AI capabilities
- Guide users to relevant solutions based on their needs
- Be friendly, professional, and knowledgeable

Key services include:
- AI Email Automation
- AI Course Creator Platform
- Generative Engine Optimisation (GEO)
- SEO Services
- Omnichannel AI Customer Support
- Voice AI Agents
- AI Sales & Lead Qualification
- Website AI Assistants
- And many more AI-powered solutions

Always be helpful and encourage visitors to explore the catalogue or book a demo if they're interested."""
        
        # Make API call to OpenAI
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": message}
            ],
            temperature=0.7,
            max_tokens=500
        )
        
        ai_response = response.choices[0].message.content
        
        return JsonResponse({
            'response': ai_response,
            'status': 'success'
        })
        
    except openai.APIError as e:
        logger.error(f'OpenAI API error: {str(e)}')
        return JsonResponse({
            'error': 'OpenAI API error',
            'response': 'I apologize, but I encountered an error processing your request. Please try again.'
        }, status=500)
    except json.JSONDecodeError as e:
        logger.error(f'JSON decode error: {str(e)}')
        return JsonResponse({
            'error': 'Invalid request format',
            'response': 'I apologize, but there was an error with your request. Please try again.'
        }, status=400)
    except Exception as e:
        logger.error(f'Chat API error: {str(e)}', exc_info=True)
        return JsonResponse({
            'error': 'Internal server error',
            'response': 'I apologize, but something went wrong. Please try again later.'
        }, status=500)

# ==================== Structured Data Extraction Endpoints ====================

@csrf_exempt
def document_upload_api(request):
    """Upload PDF document to Cloudinary"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        pdf_file = request.FILES['file']
        title = request.POST.get('title', pdf_file.name)
        
        # Validate file type
        if not pdf_file.name.lower().endswith('.pdf'):
            return JsonResponse({'error': 'Only PDF files are supported'}, status=400)
        
        # Configure Cloudinary
        cloudinary.config(
            cloud_name=settings.CLOUDINARY_CLOUD_NAME,
            api_key=settings.CLOUDINARY_API_KEY,
            api_secret=settings.CLOUDINARY_API_SECRET
        )
        
        # Upload to Cloudinary
        upload_result = cloudinary.uploader.upload(
            pdf_file,
            resource_type="raw",
            folder="extraction_documents"
        )
        file_url = upload_result['secure_url']
        
        # Create Document record
        document = Document.objects.create(title=title, file_url=file_url)
        
        return JsonResponse({
            'success': True,
            'document_id': str(document.id),
            'title': document.title,
            'file_url': document.file_url
        })
    
    except Exception as e:
        return JsonResponse({
            'error': 'Upload failed',
            'message': str(e)
        }, status=500)


@csrf_exempt
def extract_api(request):
    """Extract structured data from document using template"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        document_id = data.get('document_id')
        template_id = data.get('template_id')
        
        if not document_id or not template_id:
            return JsonResponse({'error': 'document_id and template_id are required'}, status=400)
        
        document = get_object_or_404(Document, id=document_id)
        template = get_object_or_404(ExtractionTemplate, id=template_id)
        
        # Get or create extraction run
        extraction_run, created = ExtractionRun.objects.get_or_create(
            document=document,
            template=template,
            defaults={'status': 'PENDING'}
        )
        
        if extraction_run.status == 'DONE':
            # Return existing results
            return JsonResponse({
                'success': True,
                'run_id': str(extraction_run.id),
                'extracted_json': extraction_run.extracted_json,
                'validation_json': extraction_run.validation_json,
                'status': extraction_run.status
            })
        
        # Update status to processing
        extraction_run.status = 'PROCESSING'
        extraction_run.save()
        
        try:
            # Download PDF temporarily
            import urllib.request as urllib_request
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                urllib_request.urlretrieve(document.file_url, tmp_file.name)
                tmp_path = tmp_file.name
            
            try:
                # Step 1: Extract text from PDF
                document_text = extract_text_from_pdf(tmp_path)
                
                # Step 2: Extract structured data using LLM
                template_schema = template.schema_json
                extraction_result = extract_with_llm(document_text, template_schema)
                
                extracted_data = extraction_result['extracted']
                evidence = extraction_result['evidence']
                
                # Step 3: Validate extracted data
                validation_result = validate_extraction(extracted_data, template_schema)
                
                # Store results
                extraction_run.extracted_json = {
                    'data': extracted_data,
                    'evidence': evidence
                }
                extraction_run.validation_json = validation_result
                extraction_run.status = 'DONE'
                extraction_run.save()
                
                return JsonResponse({
                    'success': True,
                    'run_id': str(extraction_run.id),
                    'extracted_json': extraction_run.extracted_json,
                    'validation_json': extraction_run.validation_json,
                    'status': extraction_run.status
                })
            
            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        
        except Exception as e:
            extraction_run.status = 'FAILED'
            extraction_run.save()
            return JsonResponse({
                'error': 'Extraction failed',
                'message': str(e)
            }, status=500)
    
    except Exception as e:
        return JsonResponse({
            'error': 'Processing failed',
            'message': str(e)
        }, status=500)


@csrf_exempt
def push_to_crm_api(request, run_id):
    """Push extracted data to local CRM demo"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        extraction_run = get_object_or_404(ExtractionRun, id=run_id)
        
        if extraction_run.status != 'DONE':
            return JsonResponse({'error': 'Extraction not completed'}, status=400)
        
        # Prepare payload
        payload = {
            'document_id': str(extraction_run.document.id),
            'document_title': extraction_run.document.title,
            'template_id': str(extraction_run.template.id),
            'template_name': extraction_run.template.name,
            'extracted_data': extraction_run.extracted_json.get('data', {}),
            'validation': extraction_run.validation_json,
            'extracted_at': extraction_run.created_at.isoformat()
        }
        
        # POST to local CRM webhook endpoint
        crm_webhook_url = request.build_absolute_uri('/api/crm/webhook/')
        response = requests.post(
            crm_webhook_url,
            json=payload,
            headers={'Content-Type': 'application/json'},
            timeout=10
        )
        
        if response.status_code in [200, 201, 202]:
            extraction_run.pushed_at = timezone.now()
            extraction_run.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Data pushed to CRM successfully',
                'pushed_at': extraction_run.pushed_at.isoformat(),
                'crm_url': '/demo/crm/'
            })
        else:
            return JsonResponse({
                'error': 'CRM webhook returned error',
                'status_code': response.status_code,
                'response': response.text
            }, status=500)
    
    except Exception as e:
        return JsonResponse({
            'error': 'Push failed',
            'message': str(e)
        }, status=500)


@csrf_exempt
def crm_webhook_api(request):
    """CRM webhook endpoint - receives pushed extraction data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Create CRM record
        crm_record = CRMRecord.objects.create(
            document_id=data.get('document_id', ''),
            document_title=data.get('document_title', ''),
            template_id=data.get('template_id', ''),
            template_name=data.get('template_name', ''),
            extracted_data=data.get('extracted_data', {}),
            validation_data=data.get('validation', {})
        )
        
        return JsonResponse({
            'success': True,
            'record_id': str(crm_record.id),
            'message': 'Record created in CRM'
        }, status=201)
    
    except Exception as e:
        return JsonResponse({
            'error': 'Webhook processing failed',
            'message': str(e)
        }, status=500)


def crm_demo(request):
    """Mini CRM demo page showing pushed extraction records"""
    records = CRMRecord.objects.all()
    return render(request, 'crm_demo.html', {'records': records})

def email_crm_demo(request):
    """Email Automation CRM demo page"""
    records = EmailCRMRecord.objects.all()
    return render(request, 'email_crm_demo.html', {'records': records})

def omnichannel_support_demo(request):
    """Omnichannel AI Customer Support demo page"""
    conversations = DemoConversation.objects.all()
    return render(request, 'omnichannel_support_demo.html', {'conversations': conversations})

def support_crm_demo(request):
    """Omnichannel Support CRM demo page"""
    tickets = DemoTicket.objects.all()
    return render(request, 'support_crm_demo.html', {'tickets': tickets})

def qualification_demo(request):
    """AI Sales Lead Qualification demo page"""
    return render(request, 'qualification_demo.html')

def qualification_leads_demo(request):
    """View all qualified leads"""
    qualifications = DemoQualification.objects.all().select_related('lead').order_by('-created_at')
    
    # Calculate stats
    hot_leads = qualifications.filter(score__gte=70).count()
    warm_leads = qualifications.filter(score__gte=40, score__lt=70).count()
    demos_booked = qualifications.filter(demo_booked=True).count()
    
    return render(request, 'qualification_leads_demo.html', {
        'qualifications': qualifications,
        'hot_leads': hot_leads,
        'warm_leads': warm_leads,
        'demos_booked': demos_booked,
    })

def automation_engine_demo(request):
    """Backend Automation Engine demo page"""
    events = DemoEvent.objects.all().order_by('-created_at')[:20]
    return render(request, 'automation_engine_demo.html', {'events': events    })

def automation_engine_demo(request):
    """Backend Automation Engine demo page"""
    events = DemoEvent.objects.all().order_by('-created_at')[:20]
    return render(request, 'automation_engine_demo.html', {'events': events})

@csrf_exempt
def run_automation_api(request):
    """Simulate an event and run workflow"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import time
        from datetime import datetime
        
        data = json.loads(request.body)
        event_type = data.get('event_type')
        
        if not event_type:
            return JsonResponse({'error': 'event_type required'}, status=400)
        
        # Create event
        event_payload = {
            'source': 'demo',
            'timestamp': datetime.now().isoformat(),
        }
        
        # Add event-specific payload
        if event_type == 'form_submission':
            event_payload.update({
                'form_name': 'Enterprise Demo Request',
                'name': 'John Smith',
                'email': 'john@enterprise.com',
                'company': 'Enterprise Corp',
            })
        elif event_type == 'instagram_dm':
            event_payload.update({
                'username': '@customer123',
                'message': 'Hi, I need help with your product',
                'channel': 'instagram',
            })
        elif event_type == 'payment_failed':
            event_payload.update({
                'customer_id': 'CUST-12345',
                'amount': 299.00,
                'reason': 'Insufficient funds',
            })
        elif event_type == 'high_score_lead':
            event_payload.update({
                'lead_name': 'Sarah Johnson',
                'lead_email': 'sarah@techcorp.com',
                'score': 85,
                'company': 'TechCorp Solutions',
            })
        elif event_type == 'support_escalation':
            event_payload.update({
                'ticket_id': 'TICKET-789',
                'customer': 'Mike Chen',
                'urgency': 'High',
                'reason': 'Technical issue',
            })
        elif event_type == 'email_received':
            event_payload.update({
                'from': 'client@example.com',
                'subject': 'Urgent: Need assistance',
                'priority': 'High',
            })
        elif event_type == 'chat_message':
            event_payload.update({
                'user': 'Visitor #1234',
                'message': 'I want to learn more about pricing',
                'channel': 'website',
            })
        
        event = DemoEvent.objects.create(
            event_type=event_type,
            payload=event_payload,
            status='processing'
        )
        
        # Simulate processing delay
        time.sleep(0.8)
        
        # Evaluate workflow rules (hardcoded for MVP)
        decision_path = []
        actions = []
        confidence_score = 95
        
        # Rule engine logic
        if event_type == 'high_score_lead':
            decision_path = [
                {'node': 'trigger', 'name': 'High Score Lead Detected', 'status': 'matched'},
                {'node': 'condition', 'name': 'Score > 80', 'status': 'matched', 'value': event_payload.get('score', 0)},
                {'node': 'condition', 'name': 'Budget Match', 'status': 'matched'},
            ]
            if event_payload.get('score', 0) >= 80:
                actions = [
                    {'action': 'assign_sales', 'target': 'Sarah (Sales)', 'status': 'completed'},
                    {'action': 'create_opportunity', 'target': event_payload.get('company', 'Unknown'), 'status': 'completed'},
                    {'action': 'send_calendar_email', 'target': event_payload.get('lead_email', ''), 'status': 'completed'},
                    {'action': 'update_crm_stage', 'target': 'Qualified', 'status': 'completed'},
                    {'action': 'slack_notification', 'target': '#sales-team', 'status': 'completed'},
                ]
            else:
                actions = [
                    {'action': 'add_to_nurture', 'target': 'Email Sequence', 'status': 'completed'},
                    {'action': 'update_crm_stage', 'target': 'Nurture', 'status': 'completed'},
                ]
        
        elif event_type == 'form_submission':
            decision_path = [
                {'node': 'trigger', 'name': 'Form Submission', 'status': 'matched'},
                {'node': 'condition', 'name': 'Form Type: Enterprise', 'status': 'matched'},
            ]
            actions = [
                {'action': 'create_lead', 'target': event_payload.get('name', ''), 'status': 'completed'},
                {'action': 'assign_sales', 'target': 'Enterprise Sales Team', 'status': 'completed'},
                {'action': 'send_confirmation_email', 'target': event_payload.get('email', ''), 'status': 'completed'},
                {'action': 'update_crm_stage', 'target': 'New Lead', 'status': 'completed'},
                {'action': 'slack_notification', 'target': '#enterprise-leads', 'status': 'completed'},
            ]
        
        elif event_type == 'instagram_dm':
            decision_path = [
                {'node': 'trigger', 'name': 'Instagram DM', 'status': 'matched'},
                {'node': 'condition', 'name': 'Channel: Instagram', 'status': 'matched'},
            ]
            actions = [
                {'action': 'create_ticket', 'target': 'Support Ticket', 'status': 'completed'},
                {'action': 'assign_team', 'target': 'Social Media Team', 'status': 'completed'},
                {'action': 'send_auto_reply', 'target': event_payload.get('username', ''), 'status': 'completed'},
                {'action': 'update_crm_stage', 'target': 'Inquiry', 'status': 'completed'},
            ]
        
        elif event_type == 'payment_failed':
            decision_path = [
                {'node': 'trigger', 'name': 'Payment Failed', 'status': 'matched'},
                {'node': 'condition', 'name': 'Retry Count < 3', 'status': 'matched'},
            ]
            actions = [
                {'action': 'send_payment_reminder', 'target': event_payload.get('customer_id', ''), 'status': 'completed'},
                {'action': 'update_subscription_status', 'target': 'Payment Failed', 'status': 'completed'},
                {'action': 'create_task', 'target': 'Follow up payment', 'status': 'completed'},
                {'action': 'slack_notification', 'target': '#billing-team', 'status': 'completed'},
            ]
        
        elif event_type == 'support_escalation':
            decision_path = [
                {'node': 'trigger', 'name': 'Support Escalation', 'status': 'matched'},
                {'node': 'condition', 'name': 'Urgency: High', 'status': 'matched'},
            ]
            actions = [
                {'action': 'escalate_ticket', 'target': 'Senior Support', 'status': 'completed'},
                {'action': 'assign_team', 'target': 'Technical Team', 'status': 'completed'},
                {'action': 'send_notification', 'target': 'Manager', 'status': 'completed'},
                {'action': 'update_priority', 'target': 'High', 'status': 'completed'},
                {'action': 'slack_notification', 'target': '#support-escalations', 'status': 'completed'},
            ]
        
        elif event_type == 'email_received':
            decision_path = [
                {'node': 'trigger', 'name': 'Email Received', 'status': 'matched'},
                {'node': 'condition', 'name': 'Priority: High', 'status': 'matched'},
            ]
            actions = [
                {'action': 'create_ticket', 'target': 'Email Ticket', 'status': 'completed'},
                {'action': 'assign_team', 'target': 'Support Team', 'status': 'completed'},
                {'action': 'send_acknowledgment', 'target': event_payload.get('from', ''), 'status': 'completed'},
                {'action': 'update_crm_stage', 'target': 'In Progress', 'status': 'completed'},
            ]
        
        elif event_type == 'chat_message':
            decision_path = [
                {'node': 'trigger', 'name': 'Chat Message', 'status': 'matched'},
                {'node': 'condition', 'name': 'Intent: Sales Inquiry', 'status': 'matched'},
            ]
            actions = [
                {'action': 'create_lead', 'target': 'Chat Lead', 'status': 'completed'},
                {'action': 'assign_sales', 'target': 'Sales Team', 'status': 'completed'},
                {'action': 'send_follow_up', 'target': 'Email Sequence', 'status': 'completed'},
                {'action': 'update_crm_stage', 'target': 'Qualified', 'status': 'completed'},
            ]
        
        # Create workflow run
        execution_time = 800  # milliseconds
        workflow_run = DemoWorkflowRun.objects.create(
            event=event,
            decision_path=decision_path,
            actions=actions,
            confidence_score=confidence_score,
            execution_time_ms=execution_time
        )
        
        # Update event status
        event.status = 'processed'
        event.processed_at = datetime.now()
        event.save()
        
        return JsonResponse({
            'success': True,
            'event_id': str(event.id),
            'workflow_run': {
                'decision_path': decision_path,
                'actions': actions,
                'confidence_score': confidence_score,
                'execution_time_ms': execution_time,
            },
            'event': {
                'type': event_type,
                'type_display': event.get_event_type_display(),
                'payload': event_payload,
                'created_at': event.created_at.isoformat(),
            }
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def process_qualification_api(request):
    """Process lead qualification"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        lead_data = data.get('lead', {})
        answers = data.get('answers', {})
        
        # Create lead
        lead = DemoLead.objects.create(
            name=lead_data.get('name', ''),
            email=lead_data.get('email', ''),
            company=lead_data.get('company', ''),
            budget_range=lead_data.get('budget_range', ''),
            service_interest=lead_data.get('service_interest', '')
        )
        
        # Calculate score
        score = 0
        score_details = []
        
        # Budget scoring
        budget = lead_data.get('budget_range', '')
        if '50k+' in budget.lower() or '100k+' in budget.lower():
            score += 30
            score_details.append({'factor': 'Budget Match', 'points': 30})
        elif '25k' in budget.lower() or '50k' in budget.lower():
            score += 20
            score_details.append({'factor': 'Budget Match', 'points': 20})
        elif '10k' in budget.lower():
            score += 10
            score_details.append({'factor': 'Budget Match', 'points': 10})
        
        # Revenue scoring
        revenue = answers.get('revenue', '').lower()
        # Extract numbers from revenue string
        import re
        revenue_numbers = re.findall(r'[\d.]+', revenue)
        revenue_value = 0
        if revenue_numbers:
            try:
                revenue_value = float(revenue_numbers[0])
                # Check for multipliers (k, m, million)
                if 'm' in revenue or 'million' in revenue:
                    revenue_value *= 1000000
                elif 'k' in revenue or 'thousand' in revenue:
                    revenue_value *= 1000
            except:
                pass
        
        if revenue_value >= 500000 or '500k+' in revenue or '1m+' in revenue or 'million' in revenue:
            score += 25
            score_details.append({'factor': 'High Revenue', 'points': 25})
        elif revenue_value >= 100000 or '100k' in revenue or '200k' in revenue or '250k' in revenue or '300k' in revenue or '400k' in revenue:
            score += 15
            score_details.append({'factor': 'Revenue', 'points': 15})
        
        # Urgency scoring
        urgency = answers.get('urgency', '').lower()
        if 'immediately' in urgency or 'asap' in urgency or 'as soon as possible' in urgency or 'now' in urgency or 'urgent' in urgency:
            score += 25
            score_details.append({'factor': 'High Urgency', 'points': 25})
        elif '30' in urgency or 'month' in urgency or 'soon' in urgency:
            score += 15
            score_details.append({'factor': 'Moderate Urgency', 'points': 15})
        elif '60' in urgency or '90' in urgency or 'quarter' in urgency:
            score += 5
            score_details.append({'factor': 'Low Urgency', 'points': 5})
        
        # CRM mismatch (needs upgrade)
        crm = answers.get('crm', '').lower()
        if 'spreadsheet' in crm or 'excel' in crm or 'none' in crm or 'nothing' in crm:
            score += 10
            score_details.append({'factor': 'CRM Upgrade Needed', 'points': 10})
        
        # Pain signals
        bottleneck = answers.get('bottleneck', '').lower()
        pain_keywords = ['manual', 'time', 'inefficient', 'slow', 'error', 'difficult', 'struggling', 'challenge']
        pain_score = sum(5 for keyword in pain_keywords if keyword in bottleneck)
        if pain_score > 0:
            score += min(pain_score, 15)
            score_details.append({'factor': 'Pain Signals', 'points': min(pain_score, 15)})
        
        # Determine intent and fit
        if score >= 70:
            intent = 'strong'
            fit_level = 'high'
            recommended_action = 'Book strategy call'
            assigned_team = 'Enterprise Sales'
            assigned_to = 'Sarah (Sales)'
            demo_booked = True
            # Simulate demo booking (30 days from now, random time)
            from datetime import datetime, timedelta
            import random
            demo_date = datetime.now() + timedelta(days=random.randint(1, 30))
            demo_date = demo_date.replace(hour=random.choice([10, 11, 14, 15, 16]), minute=0)
        elif score >= 40:
            intent = 'moderate'
            fit_level = 'medium'
            recommended_action = 'Nurture campaign'
            assigned_team = 'SMB Sales'
            assigned_to = 'Mike (Sales)'
            demo_booked = False
            demo_date = None
        else:
            intent = 'weak'
            fit_level = 'low'
            recommended_action = 'Email nurture'
            assigned_team = 'Marketing'
            assigned_to = 'Marketing Team'
            demo_booked = False
            demo_date = None
        
        # Calculate urgency days
        urgency_days = None
        if 'immediately' in urgency or 'asap' in urgency:
            urgency_days = 7
        elif '30' in urgency or 'month' in urgency:
            urgency_days = 30
        elif '60' in urgency or 'quarter' in urgency:
            urgency_days = 60
        elif '90' in urgency:
            urgency_days = 90
        
        # Create qualification
        qualification = DemoQualification.objects.create(
            lead=lead,
            answers=answers,
            score=score,
            intent=intent,
            fit_level=fit_level,
            budget_match=score >= 20 and ('50k+' in budget.lower() or '100k+' in budget.lower()),
            urgency_days=urgency_days,
            assigned_team=assigned_team,
            recommended_action=recommended_action,
            demo_booked=demo_booked,
            demo_date=demo_date,
            assigned_to=assigned_to
        )
        
        return JsonResponse({
            'success': True,
            'lead_id': str(lead.id),
            'qualification': {
                'score': score,
                'score_details': score_details,
                'intent': intent,
                'fit_level': fit_level,
                'budget_match': qualification.budget_match,
                'urgency_days': urgency_days,
                'assigned_team': assigned_team,
                'recommended_action': recommended_action,
                'demo_booked': demo_booked,
                'demo_date': demo_date.strftime('%B %d, %Y at %I:%M %p') if demo_date else None,
                'assigned_to': assigned_to
            }
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

def quality_control_demo(request):
    """Quality Control & Monitoring demo page"""
    conversations = QualityConversation.objects.all().order_by('-created_at')[:20]
    
    # Calculate global risk summary (only count analyzed conversations)
    analyzed_conversations = QualityConversation.objects.filter(risk_analysis__isnull=False)
    total_alerts = analyzed_conversations.filter(risk_analysis__alert_created=True).count()
    high_severity = analyzed_conversations.filter(risk_analysis__severity_score__gte=7.0).count()
    medium_severity = analyzed_conversations.filter(risk_analysis__severity_score__gte=4.0, risk_analysis__severity_score__lt=7.0).count()
    compliance_flags = analyzed_conversations.filter(risk_analysis__compliance_flag=True).count()
    
    return render(request, 'quality_control_demo.html', {
        'conversations': conversations,
        'total_alerts': total_alerts,
        'high_severity': high_severity,
        'medium_severity': medium_severity,
        'compliance_flags': compliance_flags,
    })

def analytics_demo(request):
    """AI Analytics & Insights demo page"""
    from datetime import datetime, timedelta
    
    # Get or create latest snapshot
    snapshot = DemoAnalyticsSnapshot.objects.first()
    if not snapshot:
        # Create default snapshot
        snapshot = DemoAnalyticsSnapshot.objects.create(
            period_start=datetime.now() - timedelta(days=30),
            period_end=datetime.now(),
            leads=1248,
            conversion_rate=4.7,
            sales_qualified_leads=214,
            deals_closed=18,
            support_tickets=62,
            churn_signals=4,
            average_response_time_minutes=192,  # 3h 12m
            channel_breakdown={
                'instagram': {'leads': 145, 'conversion': 2.1, 'qualified': 12},
                'email': {'leads': 423, 'conversion': 5.8, 'qualified': 89},
                'web': {'leads': 680, 'conversion': 6.2, 'qualified': 113},
            }
        )
    
    insights = DemoInsight.objects.filter(snapshot=snapshot).order_by('-priority', '-created_at')
    
    # Calculate executive metrics
    pipeline_value = snapshot.deals_closed * 50000  # Estimated average deal size
    revenue_this_month = snapshot.deals_closed * 45000
    
    return render(request, 'analytics_demo.html', {
        'snapshot': snapshot,
        'insights': insights,
        'pipeline_value': pipeline_value,
        'revenue_this_month': revenue_this_month,
    })

@csrf_exempt
def generate_insights_api(request):
    """Generate AI insights from analytics snapshot"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        from datetime import datetime, timedelta
        
        snapshot = DemoAnalyticsSnapshot.objects.first()
        if not snapshot:
            return JsonResponse({'error': 'No analytics snapshot found'}, status=404)
        
        # Clear existing insights
        DemoInsight.objects.filter(snapshot=snapshot).delete()
        
        insights = []
        
        # Insight 1: Conversion Drop Detection
        if snapshot.conversion_rate < 5.0:
            insights.append({
                'insight_type': 'bottleneck',
                'title': 'Conversion Rate Drop Detected',
                'description': f'Conversion rate decreased from 6.2% to {snapshot.conversion_rate}%',
                'what_happened': f'Your conversion rate has dropped from 6.2% to {snapshot.conversion_rate}% over the past month.',
                'why_it_matters': 'This represents a 24% decline in conversion efficiency. At current lead volume, this costs approximately $12,400 in lost revenue monthly.',
                'recommended_action': 'Reduce first response time to under 1 hour. High-value leads (score > 80) convert 3.4x more when responded to quickly.',
                'estimated_revenue_lift': 12400.00,
                'impact_level': 'high',
                'confidence_score': 92,
                'priority': 10,
            })
        
        # Insight 2: High-Value Lead Opportunity
        if snapshot.sales_qualified_leads > 200:
            insights.append({
                'insight_type': 'opportunity',
                'title': 'High-Value Leads Convert 3.4x Better',
                'description': 'Leads with score > 80 show significantly higher conversion rates',
                'what_happened': 'Analysis shows that leads scoring above 80 convert at 15.2% vs 4.7% average.',
                'why_it_matters': 'Increasing high-score lead volume by 30% could generate an additional $45,000 in monthly revenue.',
                'recommended_action': 'Increase paid traffic to enterprise funnel. Target keywords with higher buyer intent.',
                'estimated_revenue_lift': 45000.00,
                'impact_level': 'high',
                'confidence_score': 88,
                'priority': 9,
            })
        
        # Insight 3: Channel Performance Gap
        instagram_data = snapshot.channel_breakdown.get('instagram', {})
        if instagram_data.get('conversion', 0) < 3.0:
            insights.append({
                'insight_type': 'optimization',
                'title': 'Instagram Leads Underperforming',
                'description': f'Instagram leads convert at {instagram_data.get("conversion", 0)}% vs {snapshot.conversion_rate}% average',
                'what_happened': f'Instagram channel shows {instagram_data.get("conversion", 0)}% conversion rate, 2x lower than average.',
                'why_it_matters': 'This channel generates 145 leads monthly but only 12 qualified. Improving to average could add 6 more SQLs per month.',
                'recommended_action': 'Refine targeting or update landing page copy. Consider A/B testing Instagram ad creative.',
                'estimated_revenue_lift': 18000.00,
                'impact_level': 'medium',
                'confidence_score': 85,
                'priority': 7,
            })
        
        # Insight 4: Response Time Bottleneck
        if snapshot.average_response_time_minutes > 120:
            hours = snapshot.average_response_time_minutes // 60
            minutes = snapshot.average_response_time_minutes % 60
            insights.append({
                'insight_type': 'bottleneck',
                'title': 'Response Time Impacting Conversion',
                'description': f'Average response time is {hours}h {minutes}m, above optimal threshold',
                'what_happened': f'Current average response time is {hours}h {minutes}m. Leads responded to within 1 hour convert 2.3x better.',
                'why_it_matters': 'Reducing response time to under 1 hour could improve conversion rate by 1.2 percentage points, adding $8,500 monthly revenue.',
                'recommended_action': 'Enable auto-responses for common inquiries. Route high-score leads to dedicated team with SLA alerts.',
                'estimated_revenue_lift': 8500.00,
                'impact_level': 'high',
                'confidence_score': 90,
                'priority': 8,
            })
        
        # Insight 5: Churn Risk Alert
        if snapshot.churn_signals >= 4:
            insights.append({
                'insight_type': 'risk',
                'title': 'Churn Signals Detected',
                'description': f'{snapshot.churn_signals} customers showing churn indicators',
                'what_happened': f'AI detected {snapshot.churn_signals} customers with high churn risk based on support patterns and engagement.',
                'why_it_matters': 'At current churn rate, this could result in $22,000 monthly revenue loss if not addressed.',
                'recommended_action': 'Immediately assign account managers to high-risk customers. Offer retention incentives or feature upgrades.',
                'estimated_revenue_lift': -22000.00,  # Negative = loss prevention
                'impact_level': 'high',
                'confidence_score': 87,
                'priority': 9,
            })
        
        # Insight 6: Support Ticket Volume
        if snapshot.support_tickets > 50:
            insights.append({
                'insight_type': 'pattern',
                'title': 'Support Ticket Volume Increasing',
                'description': f'{snapshot.support_tickets} tickets this month, up from previous period',
                'what_happened': f'Support ticket volume is {snapshot.support_tickets}, indicating potential product issues or onboarding gaps.',
                'why_it_matters': 'High ticket volume can impact customer satisfaction and increase churn risk. It also diverts resources from sales.',
                'recommended_action': 'Analyze ticket categories. Create self-service resources for top 3 issues. Consider proactive onboarding improvements.',
                'estimated_revenue_lift': 0,  # Operational improvement
                'impact_level': 'medium',
                'confidence_score': 82,
                'priority': 6,
            })
        
        # Create insight records
        created_insights = []
        for insight_data in insights:
            insight = DemoInsight.objects.create(
                snapshot=snapshot,
                **insight_data
            )
            created_insights.append({
                'id': str(insight.id),
                'title': insight.title,
                'insight_type': insight.insight_type,
                'impact_level': insight.impact_level,
                'confidence_score': insight.confidence_score,
            })
        
        return JsonResponse({
            'success': True,
            'insights_generated': len(created_insights),
            'insights': created_insights
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def create_insight_task_api(request, insight_id):
    """Create a task from an insight"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        insight = get_object_or_404(DemoInsight, id=insight_id)
        
        # For MVP, just return success with a message
        # In production, this would create an actual Task object
        task_reference = f"TASK-{insight.id.hex[:8].upper()}"
        
        return JsonResponse({
            'success': True,
            'message': f'Task created from insight: {insight.title}',
            'reference': task_reference
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def draft_insight_automation_api(request, insight_id):
    """Draft an automation from an insight"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        insight = get_object_or_404(DemoInsight, id=insight_id)
        
        # For MVP, just return success with a message
        # In production, this would create an actual AutomationDraft object
        automation_reference = f"AUTO-{insight.id.hex[:8].upper()}"
        
        return JsonResponse({
            'success': True,
            'message': f'Automation draft created from insight: {insight.title}',
            'reference': automation_reference
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def notify_insight_api(request, insight_id):
    """Send notification from an insight"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        insight = get_object_or_404(DemoInsight, id=insight_id)
        
        # For MVP, just return success with a message
        # In production, this would send actual notifications (email, Slack, etc.)
        notification_reference = f"NOTIF-{insight.id.hex[:8].upper()}"
        
        return JsonResponse({
            'success': True,
            'message': f'Notification sent for insight: {insight.title}',
            'reference': notification_reference
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def analyze_quality_api(request):
    """Analyze conversation for quality and risk"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        conversation_id = data.get('conversation_id')
        
        if not conversation_id:
            return JsonResponse({'error': 'conversation_id required'}, status=400)
        
        conversation = get_object_or_404(QualityConversation, id=conversation_id)
        
        # Check if already analyzed
        if hasattr(conversation, 'risk_analysis'):
            analysis = conversation.risk_analysis
            return JsonResponse({
                'success': True,
                'analysis': {
                    'sentiment_score': analysis.sentiment_score,
                    'sentiment_label': analysis.sentiment_label,
                    'churn_risk': analysis.churn_risk,
                    'compliance_flag': analysis.compliance_flag,
                    'compliance_type': analysis.compliance_type,
                    'severity_score': analysis.severity_score,
                    'alert_created': analysis.alert_created,
                    'automation_paused': analysis.automation_paused,
                    'escalation_required': analysis.escalation_required,
                }
            })
        
        # Analyze with AI (or use hardcoded logic for MVP)
        message_lower = conversation.message.lower()
        
        # Sentiment analysis
        negative_keywords = ['cancel', 'dispute', 'refund', 'delete', 'third time', 'guarantee', 'angry', 'frustrated', 'terrible', 'worst']
        positive_keywords = ['thanks', 'thank you', 'great', 'excellent', 'appreciate', 'helpful']
        
        negative_count = sum(1 for keyword in negative_keywords if keyword in message_lower)
        positive_count = sum(1 for keyword in positive_keywords if keyword in message_lower)
        
        if negative_count > positive_count:
            sentiment_score = -0.7
            sentiment_label = 'Very Negative'
        elif negative_count > 0:
            sentiment_score = -0.4
            sentiment_label = 'Negative'
        elif positive_count > 0:
            sentiment_score = 0.6
            sentiment_label = 'Positive'
        else:
            sentiment_score = 0.0
            sentiment_label = 'Neutral'
        
        # Churn risk
        churn_keywords = ['cancel', 'leaving', 'switching', 'competitor', 'third time', 'dispute']
        churn_count = sum(1 for keyword in churn_keywords if keyword in message_lower)
        
        if churn_count >= 2 or 'cancel' in message_lower:
            churn_risk = 'high'
        elif churn_count >= 1:
            churn_risk = 'medium'
        else:
            churn_risk = 'low'
        
        # Compliance flags
        compliance_flag = False
        compliance_type = ''
        
        if 'guarantee' in message_lower and ('roi' in message_lower or 'return' in message_lower or '200%' in message_lower):
            compliance_flag = True
            compliance_type = 'ROI_Claim'
        elif 'delete' in message_lower and ('data' in message_lower or 'personal' in message_lower or 'gdpr' in message_lower):
            compliance_flag = True
            compliance_type = 'GDPR_Request'
        elif 'dispute' in message_lower or 'chargeback' in message_lower:
            compliance_flag = True
            compliance_type = 'Payment_Dispute'
        
        # Severity score (0-10)
        severity_score = 0.0
        
        # Base severity from sentiment
        if sentiment_score < -0.5:
            severity_score += 4.0
        elif sentiment_score < 0:
            severity_score += 2.0
        
        # Add for churn risk
        if churn_risk == 'high':
            severity_score += 3.0
        elif churn_risk == 'medium':
            severity_score += 1.5
        
        # Add for compliance
        if compliance_flag:
            severity_score += 2.5
        
        # Add for specific threats
        if 'dispute' in message_lower:
            severity_score += 1.5
        if 'third time' in message_lower or 'multiple' in message_lower:
            severity_score += 1.0
        
        severity_score = min(severity_score, 10.0)
        
        # Determine if escalation required
        escalation_required = severity_score >= 7.0 or compliance_flag or churn_risk == 'high'
        automation_paused = escalation_required
        alert_created = escalation_required
        
        # Create analysis
        analysis = DemoRiskAnalysis.objects.create(
            conversation=conversation,
            sentiment_score=sentiment_score,
            sentiment_label=sentiment_label,
            churn_risk=churn_risk,
            compliance_flag=compliance_flag,
            compliance_type=compliance_type,
            severity_score=severity_score,
            alert_created=alert_created,
            automation_paused=automation_paused,
            escalation_required=escalation_required
        )
        
        return JsonResponse({
            'success': True,
            'analysis': {
                'sentiment_score': sentiment_score,
                'sentiment_label': sentiment_label,
                'churn_risk': churn_risk,
                'compliance_flag': compliance_flag,
                'compliance_type': compliance_type,
                'severity_score': severity_score,
                'alert_created': alert_created,
                'automation_paused': automation_paused,
                'escalation_required': escalation_required,
            }
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def process_support_conversation_api(request):
    """Process conversation with AI - returns analysis and creates ticket"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        conversation_id = data.get('conversation_id')
        
        if not conversation_id:
            return JsonResponse({'error': 'conversation_id required'}, status=400)
        
        conversation = get_object_or_404(DemoConversation, id=conversation_id)
        
        # Check if already processed
        if conversation.is_processed and hasattr(conversation, 'ticket'):
            ticket = conversation.ticket
            return JsonResponse({
                'intent': ticket.intent,
                'urgency': ticket.urgency,
                'sentiment': ticket.sentiment,
                'assigned_team': ticket.assigned_team,
                'ticket_number': ticket.ticket_number,
                'summary': ticket.summary,
                'confidence': ticket.ai_confidence,
                'status': ticket.status,
            })
        
        # Process with AI
        client = openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        
        # Get all messages in conversation
        messages_text = conversation.initial_message
        for msg in conversation.messages.all():
            messages_text += f"\n{msg.sender}: {msg.content}"
        
        # Determine tone based on channel
        is_email = conversation.channel == 'email'
        is_social_media = conversation.channel in ['facebook', 'instagram', 'whatsapp']
        
        if is_email:
            closing_instruction = "End with 'Best regards' or similar professional closing."
            tone_instruction = "Use formal, professional email tone."
        elif is_social_media:
            closing_instruction = "Do NOT use formal closings like 'Best regards'. Keep it casual and friendly."
            tone_instruction = "Use casual, friendly social media tone."
        else:
            closing_instruction = "Keep tone conversational and friendly."
            tone_instruction = "Use conversational web chat tone."
        
        prompt = f"""Analyze this customer support conversation and provide structured response:

Channel: {conversation.get_channel_display()}
Customer: {conversation.customer_name}
Messages: {messages_text}

Important: {tone_instruction} {closing_instruction}

Provide analysis in this exact JSON format:
{{
    "intent": "one of: Billing Inquiry, Technical Support, Product Question, Complaint, Feature Request, Account Issue",
    "urgency": "one of: Low, Medium, High",
    "sentiment": "one of: Positive, Neutral, Negative",
    "assigned_team": "Team name (e.g., Finance Team, Support Team, Sales Team, etc.)",
    "summary": "Brief summary of the issue for human handoff (2-3 sentences)",
    "draft_reply": "A helpful, professional response to the customer (2-3 paragraphs). {closing_instruction}",
    "confidence": 85
}}"""

        system_message = """You are an AI customer support assistant. Analyze conversations and provide structured responses in JSON format only.

Channel-specific tone guidelines:
- Email: Use formal, professional tone. Always end with "Best regards" or similar professional closing.
- Social media (Facebook, Instagram, WhatsApp): Use casual, friendly tone. Do NOT use formal closings like "Best regards". Keep it conversational.
- Web chat: Use conversational but professional tone. No formal closings needed."""
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=800
        )
        
        content = response.choices[0].message.content.strip()
        
        # Extract JSON
        if '```json' in content:
            content = content.split('```json')[1].split('```')[0].strip()
        elif '```' in content:
            content = content.split('```')[1].split('```')[0].strip()
        
        ai_data = json.loads(content)
        
        # Generate ticket number
        import random
        ticket_number = f"TKT-{random.randint(100000, 999999)}"
        
        # Create ticket
        ticket = DemoTicket.objects.create(
            conversation=conversation,
            ticket_number=ticket_number,
            assigned_team=ai_data.get('assigned_team', 'Support Team'),
            intent=ai_data.get('intent', 'General Inquiry'),
            urgency=ai_data.get('urgency', 'Medium'),
            sentiment=ai_data.get('sentiment', 'Neutral'),
            summary=ai_data.get('summary', 'Customer inquiry requiring attention'),
            ai_confidence=ai_data.get('confidence', 85)
        )
        
        # Add AI response as message
        DemoMessage.objects.create(
            conversation=conversation,
            sender='ai',
            content=ai_data.get('draft_reply', 'Thank you for contacting us. We are looking into your inquiry.')
        )
        
        conversation.is_processed = True
        conversation.save()
        
        return JsonResponse({
            'intent': ticket.intent,
            'urgency': ticket.urgency,
            'sentiment': ticket.sentiment,
            'assigned_team': ticket.assigned_team,
            'ticket_number': ticket.ticket_number,
            'summary': ticket.summary,
            'confidence': ticket.ai_confidence,
            'status': ticket.status,
            'draft_reply': ai_data.get('draft_reply', ''),
        })
        
    except Exception as e:
        return JsonResponse({
            'error': 'Processing failed',
            'message': str(e)
        }, status=500)


def export_to_spreadsheet(request, run_id):
    """Export extracted data to XLSX spreadsheet"""
    extraction_run = get_object_or_404(ExtractionRun, id=run_id)
    
    if extraction_run.status != 'DONE':
        return JsonResponse({'error': 'Extraction not completed'}, status=400)
    
    try:
        # Get template fields
        template = extraction_run.template
        fields = template.get_fields()
        
        # Get extracted data
        extracted_data = extraction_run.extracted_json.get('data', {})
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Header row with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        headers = ['Field', 'Value', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data rows
        row_idx = 2
        for field in fields:
            field_key = field['key']
            field_label = field['label']
            value = extracted_data.get(field_key, '')
            
            # Get status from validation
            field_status = extraction_run.validation_json.get('field_status', {}).get(field_key, 'ok')
            
            ws.cell(row=row_idx, column=1, value=field_label)
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=3, value=field_status.upper())
            
            # Color code status
            if field_status == 'error':
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif field_status == 'warning':
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            else:
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            row_idx += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"extraction_{extraction_run.document.title}_{extraction_run.template.name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    except Exception as e:
        return JsonResponse({
            'error': 'Export failed',
            'message': str(e)
        }, status=500)

def crm_workflow_demo(request):
    """AI CRM Workflow Automation demo page"""
    contacts = CRMContact.objects.all().order_by('-created_at')[:50]
    tags = CRMTag.objects.all()
    
    return render(request, 'crm_workflow_demo.html', {
        'contacts': contacts,
        'tags': tags,
    })

@csrf_exempt
def simulate_crm_lead_api(request):
    """Simulate an incoming lead"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import random
        
        data = json.loads(request.body)
        lead_type = data.get('lead_type', 'website')
        
        # Pre-defined demo leads
        demo_leads = {
            'website': [
                {'name': 'Sarah Chen', 'email': 'sarah.chen@techcorp.com', 'company': 'TechCorp Inc', 'message': 'Need pricing for 50 seats, looking to implement this week. Can we schedule a call?', 'intent': 'Sales', 'urgency': 'high', 'fit_score': 85},
                {'name': 'Michael Rodriguez', 'email': 'mrodriguez@startup.io', 'company': 'Startup.io', 'message': 'Just browsing, send me some info about your services', 'intent': 'Sales', 'urgency': 'low', 'fit_score': 45},
                {'name': 'Emily Watson', 'email': 'ewatson@enterprise.com', 'company': 'Enterprise Solutions', 'message': 'Partnership inquiry - we work with 500+ SMBs and think there could be synergy', 'intent': 'Partnership', 'urgency': 'medium', 'fit_score': 70},
            ],
            'instagram': [
                {'name': 'Jessica Park', 'email': 'jpark@design.co', 'company': 'Design Co', 'message': 'Love your product! How do I get started?', 'intent': 'Sales', 'urgency': 'medium', 'fit_score': 60},
                {'name': 'David Kim', 'email': 'dkim@agency.com', 'company': 'Digital Agency', 'message': 'Interested in your automation features', 'intent': 'Sales', 'urgency': 'low', 'fit_score': 55},
            ],
            'email': [
                {'name': 'Robert Thompson', 'email': 'rthompson@company.com', 'company': 'Company Inc', 'message': 'Support issue: login not working. Need help ASAP', 'intent': 'Support', 'urgency': 'high', 'fit_score': 0},
                {'name': 'Lisa Anderson', 'email': 'landerson@business.com', 'company': 'Business Solutions', 'message': 'Refund request - service not as described', 'intent': 'Billing', 'urgency': 'high', 'fit_score': 0},
            ],
            'referral': [
                {'name': 'James Wilson', 'email': 'jwilson@referral.com', 'company': 'Referral Corp', 'message': 'Referred by TechCorp - interested in enterprise plan', 'intent': 'Sales', 'urgency': 'medium', 'fit_score': 75},
            ]
        }
        
        # Select random lead from type
        leads = demo_leads.get(lead_type, demo_leads['website'])
        lead_data = random.choice(leads)
        
        # Determine route and pipeline stage
        if lead_data['intent'] == 'Sales':
            if lead_data['fit_score'] >= 70:
                route = 'Sales'
                stage = 'New Lead'
                assigned = 'Sales - Sky'
            else:
                route = 'Sales'
                stage = 'New Lead'
                assigned = 'Sales - Nurture Team'
        elif lead_data['intent'] == 'Support':
            route = 'Support'
            stage = 'Support Queue'
            assigned = 'Support - Alex'
        elif lead_data['intent'] == 'Billing':
            route = 'Ops'
            stage = 'Billing Review'
            assigned = 'Ops - Finance'
        else:
            route = 'Partnerships'
            stage = 'Partnership Inquiry'
            assigned = 'Partnerships - Team'
        
        # Create contact
        contact = CRMContact.objects.create(
            name=lead_data['name'],
            email=lead_data['email'],
            company=lead_data['company'],
            channel=lead_type,
            source_message=lead_data['message'],
            intent=lead_data['intent'],
            urgency=lead_data['urgency'],
            fit_score=lead_data['fit_score'],
            recommended_route=route,
            pipeline_stage=stage,
            assigned_to=assigned
        )
        
        # Create activity
        CRMActivity.objects.create(
            contact=contact,
            activity_type='created',
            description=f'Contact created from {contact.get_channel_display()}'
        )
        
        return JsonResponse({
            'success': True,
            'contact': {
                'id': str(contact.id),
                'name': contact.name,
                'email': contact.email,
                'company': contact.company,
                'channel': contact.channel,
                'message': contact.source_message,
                'intent': contact.intent,
                'urgency': contact.urgency,
                'fit_score': contact.fit_score,
                'route': contact.recommended_route,
                'stage': contact.pipeline_stage,
                'assigned_to': contact.assigned_to,
                'created_at': contact.created_at.isoformat(),
            }
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def get_crm_contact_api(request, contact_id):
    """Get contact details"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        contact = get_object_or_404(CRMContact, id=contact_id)
        activities = CRMActivity.objects.filter(contact=contact).order_by('-created_at')[:20]
        tasks = CRMTask.objects.filter(contact=contact).order_by('-created_at')
        tickets = CRMTicket.objects.filter(contact=contact).order_by('-created_at')
        tags = contact.tags.all()
        
        return JsonResponse({
            'success': True,
            'contact': {
                'id': str(contact.id),
                'name': contact.name,
                'email': contact.email,
                'company': contact.company,
                'channel': contact.channel,
                'intent': contact.intent,
                'urgency': contact.urgency,
                'fit_score': contact.fit_score,
                'pipeline_stage': contact.pipeline_stage,
                'assigned_to': contact.assigned_to,
                'source_message': contact.source_message,
            },
            'activities': [
                {
                    'type': act.activity_type,
                    'description': act.description,
                    'created_at': act.created_at.isoformat(),
                } for act in activities
            ],
            'tasks': [
                {
                    'id': str(task.id),
                    'title': task.title,
                    'priority': task.priority,
                    'assigned_to': task.assigned_to,
                    'completed': task.completed,
                    'created_at': task.created_at.isoformat(),
                } for task in tasks
            ],
            'tickets': [
                {
                    'id': str(ticket.id),
                    'title': ticket.title,
                    'status': ticket.status,
                    'assigned_to': ticket.assigned_to,
                    'created_at': ticket.created_at.isoformat(),
                } for ticket in tickets
            ],
            'tags': [
                {
                    'name': tag.name,
                    'color': tag.color,
                } for tag in tags
            ],
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
def run_crm_automation_api(request, contact_id):
    """Run automation for a contact"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import time
        start_time = time.time()
        
        contact = get_object_or_404(CRMContact, id=contact_id)
        data = json.loads(request.body)
        auto_mode = data.get('auto_mode', True)
        
        rules_matched = []
        actions_executed = []
        
        # Rule 1: Tag based on channel
        if contact.channel == 'website':
            tag_name = 'Inbound / Website'
            rules_matched.append('Tag by Channel')
        elif contact.channel == 'instagram':
            tag_name = 'Inbound / Instagram'
            rules_matched.append('Tag by Channel')
        elif contact.channel == 'email':
            tag_name = 'Inbound / Email'
            rules_matched.append('Tag by Channel')
        elif contact.channel == 'referral':
            tag_name = 'Referral'
            rules_matched.append('Tag by Channel')
        else:
            tag_name = 'Inbound'
            rules_matched.append('Tag by Channel')
        
        # Create or get tag
        tag, _ = CRMTag.objects.get_or_create(name=tag_name, defaults={'color': 'blue'})
        contact.tags.add(tag)
        actions_executed.append(f'Tagged: {tag_name}')
        
        # Create activity
        CRMActivity.objects.create(
            contact=contact,
            activity_type='tagged',
            description=f'Tagged: {tag_name}'
        )
        
        # Rule 2: Set pipeline stage
        if contact.pipeline_stage:
            rules_matched.append('Set Pipeline Stage')
            actions_executed.append(f'Set Pipeline Stage: {contact.pipeline_stage}')
            CRMActivity.objects.create(
                contact=contact,
                activity_type='stage_changed',
                description=f'Stage set to: {contact.pipeline_stage}'
            )
        
        # Rule 3: Assign owner
        if contact.assigned_to:
            rules_matched.append('Assign Owner')
            actions_executed.append(f'Assigned to: {contact.assigned_to}')
            CRMActivity.objects.create(
                contact=contact,
                activity_type='assigned',
                description=f'Assigned to: {contact.assigned_to}'
            )
        
        # Rule 4: Create task based on intent and urgency
        if contact.intent == 'Sales' and contact.urgency == 'high':
            task = CRMTask.objects.create(
                contact=contact,
                title='Call within 15 minutes',
                description=f'High-intent lead: {contact.source_message}',
                priority='urgent',
                assigned_to=contact.assigned_to,
                due_date=timezone.now() + timedelta(minutes=15)
            )
            rules_matched.append('Create Urgent Task')
            actions_executed.append(f'Task created: {task.title}')
            CRMActivity.objects.create(
                contact=contact,
                activity_type='task_created',
                description=f'Task: {task.title}'
            )
        elif contact.intent == 'Sales' and contact.fit_score >= 70:
            task = CRMTask.objects.create(
                contact=contact,
                title='Schedule discovery call',
                description=f'High-fit lead from {contact.company}',
                priority='high',
                assigned_to=contact.assigned_to,
                due_date=timezone.now() + timedelta(hours=4)
            )
            rules_matched.append('Create High-Priority Task')
            actions_executed.append(f'Task created: {task.title}')
            CRMActivity.objects.create(
                contact=contact,
                activity_type='task_created',
                description=f'Task: {task.title}'
            )
        elif contact.intent == 'Support':
            ticket = CRMTicket.objects.create(
                contact=contact,
                title=f'Support: {contact.source_message[:50]}',
                description=contact.source_message,
                status='open',
                assigned_to=contact.assigned_to
            )
            rules_matched.append('Create Support Ticket')
            actions_executed.append(f'Ticket created: {ticket.title}')
            CRMActivity.objects.create(
                contact=contact,
                activity_type='ticket_created',
                description=f'Ticket: {ticket.title}'
            )
        
        # Rule 5: Schedule follow-up sequence for sales leads
        if contact.intent == 'Sales' and contact.fit_score >= 50:
            rules_matched.append('Schedule Follow-up Sequence')
            actions_executed.append('Follow-up sequence: Day 1, Day 3, Day 7')
            CRMActivity.objects.create(
                contact=contact,
                activity_type='followup_scheduled',
                description='Follow-up sequence started: Day 1, Day 3, Day 7'
            )
        
        # Calculate confidence
        confidence = min(95, 70 + (contact.fit_score // 2))
        
        # Create automation run log
        execution_time = int((time.time() - start_time) * 1000)
        automation_run = CRMAutomationRun.objects.create(
            contact=contact,
            triggered_by=contact.channel,
            rules_matched=rules_matched,
            actions_executed=actions_executed,
            confidence_score=confidence,
            execution_time_ms=execution_time
        )
        
        CRMActivity.objects.create(
            contact=contact,
            activity_type='automation_triggered',
            description=f'Automation executed: {len(actions_executed)} actions'
        )
        
        # Get updated contact data
        activities = CRMActivity.objects.filter(contact=contact).order_by('-created_at')
        tasks = CRMTask.objects.filter(contact=contact).order_by('-created_at')
        tickets = CRMTicket.objects.filter(contact=contact).order_by('-created_at')
        tags = contact.tags.all()
        
        return JsonResponse({
            'success': True,
            'automation': {
                'id': str(automation_run.id),
                'rules_matched': rules_matched,
                'actions_executed': actions_executed,
                'confidence_score': confidence,
                'execution_time_ms': execution_time,
            },
            'contact': {
                'id': str(contact.id),
                'name': contact.name,
                'email': contact.email,
                'company': contact.company,
                'channel': contact.channel,
                'intent': contact.intent,
                'urgency': contact.urgency,
                'fit_score': contact.fit_score,
                'pipeline_stage': contact.pipeline_stage,
                'assigned_to': contact.assigned_to,
                'tags': [{'name': tag.name, 'color': tag.color} for tag in tags],
            },
            'activities': [
                {
                    'type': act.activity_type,
                    'description': act.description,
                    'created_at': act.created_at.isoformat(),
                } for act in activities[:20]
            ],
            'tasks': [
                {
                    'id': str(task.id),
                    'title': task.title,
                    'priority': task.priority,
                    'assigned_to': task.assigned_to,
                    'completed': task.completed,
                    'created_at': task.created_at.isoformat(),
                } for task in tasks
            ],
            'tickets': [
                {
                    'id': str(ticket.id),
                    'title': ticket.title,
                    'status': ticket.status,
                    'assigned_to': ticket.assigned_to,
                    'created_at': ticket.created_at.isoformat(),
                } for ticket in tickets
            ],
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)